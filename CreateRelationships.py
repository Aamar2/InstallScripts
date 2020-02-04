#to generate Multibyte Data
import openpyxl
from openpyxl import load_workbook
import requests,os,sys,time

def updateExcel(path,rows):
	wb = load_workbook(path)
	ws1 = wb.get_sheet_by_name("Terms")
	terms=''
	relate=''
	relation=''

	for i in range(2,ws1.max_row+1):
		terms=terms+str(ws1.cell(row=i,column=2).value)+','+'\n'
		relate="related"+','+'\n'+relate
		relation="Relationships"+','+'\n'+relation
	terms=terms[:len(terms)-1]
	relate=relate[:len(relate)-1]
	relation=relation[:len(relation)-1]
	print(terms)
	print(relate)
	print(relation)
		
	for rowNum in range(ws1.max_row+1,ws1.max_row+rows):
		for columnNum in range(1,ws1.max_column+1):
			if(ws1.cell(row=1,column=columnNum).value == 'Term Name'):
				try:
					ws1.cell(row=rowNum,column=columnNum).value = 'Term' + str(rowNum-1)
				except IndexError:
					print("error occurred")
					ws1.cell(row=rowNum,column=1).value = ''
					wb.save(path)
					break
			elif(ws1.cell(row=1,column=columnNum).value == 'Related Term Names'):
				try:	
					ws1.cell(row=rowNum,column=columnNum).value = terms
				except IndexError:
					print("error occurred")
					ws1.cell(row=rowNum,column=columnNum).value = ''
					wb.save(path)
					break
			elif(ws1.cell(row=1,column=columnNum).value == 'Relationship Type Name'):
				try:	
					ws1.cell(row=rowNum,column=columnNum).value = relate
				except IndexError:
					print("error occurred")
					ws1.cell(row=rowNum,column=columnNum).value = ''
					wb.save(path)
					break
			elif(ws1.cell(row=1,column=columnNum).value == 'Related Glossary Name'):
				try:	
					ws1.cell(row=rowNum,column=columnNum).value = relation
				except IndexError:
					print("error occurred")
					ws1.cell(row=rowNum,column=columnNum).value = ''
					wb.save(path)
					break							
			else:
				desc = ws1.cell(row=i,column=columnNum).value
				ws1.cell(row=rowNum,column=columnNum).value = desc
		terms=terms+','+'\n'+str(ws1.cell(row=rowNum-1,column=2).value)
		relate=relate+','+'\n'+"related"
		relation=relation+','+'\n'+"Relationships"
		time.sleep(.001)	
		
			
	wb.save(path)

	
result = updateExcel("C:\\python_files\\Excel.xlsx",100)

print(result)
	