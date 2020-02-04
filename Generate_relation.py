#to generate Multibyte Data
import openpyxl
from openpyxl import load_workbook
import requests,os,sys,time

def updateExcel(path,rows):
	wb = load_workbook(path)
	ws1 = wb.get_sheet_by_name("Terms")
		
	for rowNum in range(ws1.max_row+1,ws1.max_row+rows):
		for columnNum in range(1,ws1.max_column+1):
			if(ws1.cell(row=1,column=columnNum).value == 'Term Name'):
				try:
					ws1.cell(row=rowNum,column=columnNum).value = 'Term' + str(rowNum-1)
				except IndexError:
					print("error occurred")
					ws1.cell(row=rowNum,column=1).value = ''
					wb.save(path)
					break
			elif(ws1.cell(row=1,column=columnNum).value == 'Related Term Names'):
				try:	
					ws1.cell(row=rowNum,column=columnNum).value = ws1.cell(row=rowNum-1,column=2).value
				except IndexError:
					print("error occurred")
					ws1.cell(row=rowNum,column=columnNum).value = ''
					wb.save(path)
					break
			elif(ws1.cell(row=1,column=columnNum).value == 'Relationship Type Name'):
				try:	
					ws1.cell(row=rowNum,column=columnNum).value = "related"
				except IndexError:
					print("error occurred")
					ws1.cell(row=rowNum,column=columnNum).value = ''
					wb.save(path)
					break
			elif(ws1.cell(row=1,column=columnNum).value == 'Related Glossary Name'):
				try:	
					ws1.cell(row=rowNum,column=columnNum).value = "Relationships"
				except IndexError:
					print("error occurred")
					ws1.cell(row=rowNum,column=columnNum).value = ''
					wb.save(path)
					break							
			else:
				desc = ws1.cell(row=2,column=columnNum).value
				ws1.cell(row=rowNum,column=columnNum).value = desc
		time.sleep(.001)	
		
			
	wb.save(path)

	
result = updateExcel("C:\\python_files\\Export_Rel.xlsx",100)

print(result)
	