#to generate Multibyte Data
import openpyxl
from openpyxl import load_workbook
from faker import Factory
import requests,os,sys,time


def generateLocale(locale,noOfChars, cat = 'name'):
	myGenerator = Factory.create(locale)
	x = ''
	for i in range(5):
		if(cat=='name'):
			x = x.join(myGenerator.simple_profile().get('name').replace('-','').replace(' ',''))
		else:	
			x = x.join(myGenerator.paragraphs())
	return x[:noOfChars].lower()

def updateExcel(path,rows,locale):
	wb = load_workbook(path)
	ws1 = wb.get_sheet_by_name("Terms")
	i = 2
	j = 10
	k = 1
	for rowNum in range(ws1.max_row+1,ws1.max_row+rows):
		for columnNum in range(1,ws1.max_column+1):
			if(ws1.cell(row=1,column=columnNum).value == 'Term Name'):
				try:
					ws1.cell(row=rowNum,column=columnNum).value = generateLocale(locale,j)
				except IndexError:
					print("error occurred")
					ws1.cell(row=rowNum,column=1).value = ''
					wb.save(path)
					break
			elif(ws1.cell(row=1,column=columnNum).value == 'Description'):
				try:	
					ws1.cell(row=rowNum,column=columnNum).value = generateLocale(locale,k,'desc')
				except IndexError:
					print("error occurred")
					ws1.cell(row=rowNum,column=columnNum).value = ''
					wb.save(path)
					break	
			else:
				desc = ws1.cell(row=i,column=columnNum).value
				ws1.cell(row=rowNum,column=columnNum).value = desc
		time.sleep(.001)	
		i = i+1
		j = j+1
		k = k+1
		if(j == 200):
			j = 10
			
	wb.save(path)
	return j
	
result = updateExcel("C:\\BG_TESTDATA\\Japanesse\\ArabicData.xlsx",100,'ar_AR')

print(result)
	